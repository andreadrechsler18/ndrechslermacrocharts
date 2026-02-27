"""
Fetch Federal Reserve regional survey diffusion indexes.

Sources: Philadelphia, Dallas, Richmond, New York, Kansas City
Each city has manufacturing and services surveys.
"""

import csv
import io
import os
import sys
from datetime import datetime, timezone

import openpyxl
import requests

sys.path.insert(0, os.path.dirname(__file__))
from utils import write_json, retry_request

# ---------------------------------------------------------------------------
# Download URLs
# ---------------------------------------------------------------------------
SOURCES = {
    # Philadelphia
    "philly_mfg": {
        "url": "https://www.philadelphiafed.org/-/media/FRBP/Assets/Surveys-And-Data/MBOS/Historical-Data/Diffusion-Indexes/bos_dif.csv",
        "format": "csv",
    },
    "philly_svc": {
        "url": "https://www.philadelphiafed.org/-/media/FRBP/Assets/Surveys-And-Data/NBOS/nboshistory.xlsx",
        "format": "xlsx",
    },
    # Dallas
    "dallas_mfg": {
        "url": "https://www.dallasfed.org/~/media/Documents/research/surveys/tmos/documents/index_sa.xls",
        "format": "xlsx",  # despite .xls extension, it's actually xlsx
    },
    "dallas_svc": {
        "url": "https://www.dallasfed.org/~/media/Documents/research/surveys/tssos/documents/tssos_index_sa.xls",
        "format": "xlsx",
    },
    # Richmond
    "richmond_mfg": {
        "url": "https://www.richmondfed.org/-/media/RichmondFedOrg/region_communities/regional_data_analysis/regional_economy/surveys_of_business_conditions/manufacturing/data/mfg_historicaldata.xlsx",
        "format": "xlsx",
    },
    "richmond_svc": {
        "url": "https://www.richmondfed.org/-/media/RichmondFedOrg/region_communities/regional_data_analysis/regional_economy/surveys_of_business_conditions/non-manufacturing/data/nmf_historicaldata.xlsx",
        "format": "xlsx",
    },
    # New York
    "ny_mfg": {
        "url": "https://www.newyorkfed.org/medialibrary/media/Survey/Empire/data/ESMS_SeasonallyAdjusted_Diffusion.csv",
        "format": "csv",
    },
    "ny_svc": {
        "url": "https://www.newyorkfed.org/medialibrary/media/survey/business_leaders/data/bls_notseasonallyadjusted_diffusion.csv",
        "format": "csv",
    },
    # Kansas City
    "kc_mfg": {
        "url": "https://www.kansascityfed.org/Manufacturing/documents/14937/2026Feb26historicalmfg.xlsx",
        "format": "kc_xlsx",
    },
    "kc_svc": {
        "url": "https://www.kansascityfed.org/Services/documents/14259/2026Janhistoricalserv.xlsx",
        "format": "kc_xlsx",
    },
}

# ---------------------------------------------------------------------------
# Column mappings: source column -> (standardized_code, display_name)
# Suffix _C = current, _F = future (6-month ahead expectations)
# ---------------------------------------------------------------------------

PHILLY_MFG_COLS = {
    # Current
    "GAC": ("GA_C", "General Activity"),
    "NOC": ("NO_C", "New Orders"),
    "SHC": ("SH_C", "Shipments"),
    "UOC": ("UO_C", "Unfilled Orders"),
    "DTC": ("DT_C", "Delivery Times"),
    "IVC": ("IV_C", "Inventories"),
    "PPC": ("PP_C", "Prices Paid"),
    "PRC": ("PR_C", "Prices Received"),
    "NEC": ("EMP_C", "Employment"),
    "AWC": ("AW_C", "Avg Workweek"),
    # Future
    "GAF": ("GA_F", "Future General Activity"),
    "NOF": ("NO_F", "Future New Orders"),
    "SHF": ("SH_F", "Future Shipments"),
    "UOF": ("UO_F", "Future Unfilled Orders"),
    "DTF": ("DT_F", "Future Delivery Times"),
    "IVF": ("IV_F", "Future Inventories"),
    "PPF": ("PP_F", "Future Prices Paid"),
    "PRF": ("PR_F", "Future Prices Received"),
    "NEF": ("EMP_F", "Future Employment"),
    "AWF": ("AW_F", "Future Avg Workweek"),
    "CEF": ("CE_F", "Future Capital Expenditures"),
}

PHILLY_SVC_COLS = {
    # Current (from Diffusion sheet, columns are like garbndif_sa)
    "garbndif_sa": ("GA_C", "General Activity (Regional)"),
    "gabndif_sa": ("GAC_C", "General Activity (Company)"),
    "nobndif_sa": ("NO_C", "New Orders"),
    "srbndif_sa": ("REV_C", "Sales/Revenue"),
    "uobndif_sa": ("UO_C", "Unfilled Orders"),
    "ivbndif_sa": ("IV_C", "Inventories"),
    "ppbndif_sa": ("PP_C", "Prices Paid"),
    "prbndif_sa": ("PR_C", "Prices Received"),
    "nfbndif_sa": ("EMP_C", "Full-Time Employment"),
    "npbndif_sa": ("PEMP_C", "Part-Time Employment"),
    "awbndif_sa": ("AW_C", "Avg Workweek"),
    "wbbndif_sa": ("WB_C", "Wages & Benefits"),
    "cpbndif_sa": ("CEP_C", "CapEx (Plant)"),
    "cebndif_sa": ("CEE_C", "CapEx (Equipment)"),
    # Future
    "garfbndif_sa": ("GA_F", "Future General Activity (Regional)"),
    "gafbndif_sa": ("GAC_F", "Future General Activity (Company)"),
}

DALLAS_MFG_COLS = {
    # Current
    "Prod": ("PROD_C", "Production"),
    "Capu": ("CU_C", "Capacity Utilization"),
    "Vnwo": ("NO_C", "New Orders"),
    "Gro": ("GA_C", "General Business Activity"),
    "Ufil": ("UO_C", "Unfilled Orders"),
    "Vshp": ("SH_C", "Shipments"),
    "Dtm": ("DT_C", "Delivery Times"),
    "Fgi": ("IV_C", "Finished Goods Inventories"),
    "Prm": ("PP_C", "Prices Paid (Raw Materials)"),
    "Pfg": ("PR_C", "Prices Received (Finished Goods)"),
    "Wgs": ("WB_C", "Wages & Benefits"),
    "Nemp": ("EMP_C", "Employment"),
    "Avgwk": ("AW_C", "Hours Worked"),
    "Cexp": ("CE_C", "Capital Expenditures"),
    "Colk": ("OL_C", "Company Outlook"),
    "Bact": ("BA_C", "Business Activity"),
    "Uncr": ("UN_C", "Outlook Uncertainty"),
    # Future
    "Fprod": ("PROD_F", "Future Production"),
    "Fcapu": ("CU_F", "Future Capacity Utilization"),
    "Fvnwo": ("NO_F", "Future New Orders"),
    "Fgro": ("GA_F", "Future General Business Activity"),
    "Fufil": ("UO_F", "Future Unfilled Orders"),
    "Fvshp": ("SH_F", "Future Shipments"),
    "Fdtm": ("DT_F", "Future Delivery Times"),
    "Ffgi": ("IV_F", "Future Finished Goods Inventories"),
    "Fprm": ("PP_F", "Future Prices Paid"),
    "Fpfg": ("PR_F", "Future Prices Received"),
    "Fwgs": ("WB_F", "Future Wages & Benefits"),
    "Fnemp": ("EMP_F", "Future Employment"),
    "Favgwk": ("AW_F", "Future Hours Worked"),
    "Fcexp": ("CE_F", "Future Capital Expenditures"),
    "Fcolk": ("OL_F", "Future Company Outlook"),
    "Fbact": ("BA_F", "Future Business Activity"),
}

DALLAS_SVC_COLS = {
    # Current
    "rev": ("REV_C", "Revenue"),
    "emp": ("EMP_C", "Employment"),
    "pemp": ("PEMP_C", "Part-Time Employment"),
    "avgwk": ("AW_C", "Hours Worked"),
    "wgs": ("WB_C", "Wages & Benefits"),
    "inp": ("PP_C", "Input Prices"),
    "sell": ("PR_C", "Selling Prices"),
    "cexp": ("CE_C", "Capital Expenditures"),
    "colk": ("OL_C", "Company Outlook"),
    "bact": ("GA_C", "General Business Activity"),
    "uncr": ("UN_C", "Outlook Uncertainty"),
    # Future
    "frev": ("REV_F", "Future Revenue"),
    "femp": ("EMP_F", "Future Employment"),
    "fpemp": ("PEMP_F", "Future Part-Time Employment"),
    "favgwk": ("AW_F", "Future Hours Worked"),
    "fwgs": ("WB_F", "Future Wages & Benefits"),
    "finp": ("PP_F", "Future Input Prices"),
    "fsell": ("PR_F", "Future Selling Prices"),
    "fcexp": ("CE_F", "Future Capital Expenditures"),
    "fcolk": ("OL_F", "Future Company Outlook"),
    "fbact": ("GA_F", "Future General Business Activity"),
}

# Richmond: SA columns start with sa_mfg_ or sa_svc_
# _c suffix = current, _e suffix = expected/future
RICHMOND_MFG_COLS = {
    "sa_mfg_composite": ("GA_C", "Composite Index"),
    "sa_mfg_ship_c": ("SH_C", "Shipments"),
    "sa_mfg_new_orders_c": ("NO_C", "New Orders"),
    "sa_mfg_bk_logs_c": ("UO_C", "Backlogs"),
    "sa_mfg_cap_util_c": ("CU_C", "Capacity Utilization"),
    "sa_mfg_vend_lead_c": ("DT_C", "Vendor Lead Time"),
    "sa_mfg_emp_c": ("EMP_C", "Employment"),
    "sa_mfg_workwk_c": ("AW_C", "Avg Workweek"),
    "sa_mfg_wage_c": ("WB_C", "Wages"),
    "sa_mfg_fd_gds_inv_c": ("IV_C", "Finished Goods Inventories"),
    "sa_mfg_raw_mats_inv_c": ("RMI_C", "Raw Materials Inventories"),
    "sa_mfg_local_bus_cond_c": ("LBC_C", "Local Business Conditions"),
    "sa_mfg_capital_expnd_c": ("CE_C", "Capital Expenditures"),
    "sa_mfg_equip_sftw_expnd_c": ("ESE_C", "Equipment & Software Spending"),
    "sa_mfg_bus_svcs_expnd_c": ("BSE_C", "Business Services Spending"),
    "sa_mfg_nec_skls_avail_c": ("SKL_C", "Skilled Workers Availability"),
    # Future/Expected
    "sa_mfg_ship_e": ("SH_F", "Future Shipments"),
    "sa_mfg_new_orders_e": ("NO_F", "Future New Orders"),
    "sa_mfg_bk_logs_e": ("UO_F", "Future Backlogs"),
    "sa_mfg_cap_util_e": ("CU_F", "Future Capacity Utilization"),
    "sa_mfg_vend_lead_e": ("DT_F", "Future Vendor Lead Time"),
    "sa_mfg_emp_e": ("EMP_F", "Future Employment"),
    "sa_mfg_workwk_e": ("AW_F", "Future Avg Workweek"),
    "sa_mfg_wage_e": ("WB_F", "Future Wages"),
    "sa_mfg_fd_gds_inv_e": ("IV_F", "Future Finished Goods Inventories"),
    "sa_mfg_raw_mats_inv_e": ("RMI_F", "Future Raw Materials Inventories"),
    "sa_mfg_local_bus_cond_e": ("LBC_F", "Future Local Business Conditions"),
    "sa_mfg_capital_expnd_e": ("CE_F", "Future Capital Expenditures"),
    "sa_mfg_nec_skls_avail_e": ("SKL_F", "Future Skilled Workers Availability"),
}

RICHMOND_SVC_COLS = {
    "sa_svc_revs_sales_c": ("REV_C", "Revenue/Sales"),
    "sa_svc_emp_c": ("EMP_C", "Employment"),
    "sa_svc_ave_wage_c": ("WB_C", "Wages"),
    "sa_svc_demand_c": ("DEM_C", "Demand"),
    "sa_svc_local_bus_cond_c": ("LBC_C", "Local Business Conditions"),
    "sa_svc_nec_skls_avail_c": ("SKL_C", "Skilled Workers Availability"),
    "sa_svc_ave_workwk_c": ("AW_C", "Avg Workweek"),
    "sa_svc_capital_expnd_c": ("CE_C", "Capital Expenditures"),
    "sa_svc_equip_sftw_expnd_c": ("ESE_C", "Equipment & Software Spending"),
    "sa_svc_bus_svcs_expnd_c": ("BSE_C", "Business Services Spending"),
    # Future/Expected
    "sa_svc_revs_sales_e": ("REV_F", "Future Revenue/Sales"),
    "sa_svc_emp_e": ("EMP_F", "Future Employment"),
    "sa_svc_ave_wage_e": ("WB_F", "Future Wages"),
    "sa_svc_demand_e": ("DEM_F", "Future Demand"),
    "sa_svc_local_bus_cond_e": ("LBC_F", "Future Local Business Conditions"),
    "sa_svc_nec_skls_avail_e": ("SKL_F", "Future Skilled Workers Availability"),
    "sa_svc_ave_workwk_e": ("AW_F", "Future Avg Workweek"),
    "sa_svc_capital_expnd_e": ("CE_F", "Future Capital Expenditures"),
    "sa_svc_equip_sftw_expnd_e": ("ESE_F", "Future Equipment & Software Spending"),
    "sa_svc_bus_svcs_expnd_e": ("BSE_F", "Future Business Services Spending"),
}

# NY Manufacturing: columns like GACDISA (current SA), GAFDISA (future SA)
NY_MFG_COLS = {
    "GACDISA": ("GA_C", "General Business Conditions"),
    "NOCDISA": ("NO_C", "New Orders"),
    "SHCDISA": ("SH_C", "Shipments"),
    "UOCDISA": ("UO_C", "Unfilled Orders"),
    "DTCDISA": ("DT_C", "Delivery Times"),
    "IVCDISA": ("IV_C", "Inventories"),
    "PPCDISA": ("PP_C", "Prices Paid"),
    "PRCDISA": ("PR_C", "Prices Received"),
    "NECDISA": ("EMP_C", "Number of Employees"),
    "AWCDISA": ("AW_C", "Avg Employee Workweek"),
    "ASCDISA": ("TS_C", "Technology Spending"),
    # Future
    "GAFDISA": ("GA_F", "Future General Business Conditions"),
    "NOFDISA": ("NO_F", "Future New Orders"),
    "SHFDISA": ("SH_F", "Future Shipments"),
    "UOFDISA": ("UO_F", "Future Unfilled Orders"),
    "DTFDISA": ("DT_F", "Future Delivery Times"),
    "IVFDISA": ("IV_F", "Future Inventories"),
    "PPFDISA": ("PP_F", "Future Prices Paid"),
    "PRFDISA": ("PR_F", "Future Prices Received"),
    "NEFDISA": ("EMP_F", "Future Number of Employees"),
    "AWFDISA": ("AW_F", "Future Avg Employee Workweek"),
    "CEFDISA": ("CE_F", "Future Capital Expenditures"),
    "ASFDISA": ("TS_F", "Future Technology Spending"),
}

# NY Services (NOT seasonally adjusted): columns like BACDINA, BAFDINA
NY_SVC_COLS = {
    "BACDINA": ("GA_C", "Business Activity"),
    "BCCDINA": ("BC_C", "Business Climate"),
    "EMCDINA": ("EMP_C", "Employment"),
    "WPCDINA": ("WB_C", "Wages"),
    "PPCDINA": ("PP_C", "Prices Paid"),
    "PRCDINA": ("PR_C", "Prices Received"),
    "CSCDINA": ("CE_C", "Capital Spending"),
    "ASCDINA": ("TS_C", "Technology Spending"),
    # Future
    "BAFDINA": ("GA_F", "Future Business Activity"),
    "BCFDINA": ("BC_F", "Future Business Climate"),
    "EMFDINA": ("EMP_F", "Future Employment"),
    "WPFDINA": ("WB_F", "Future Wages"),
    "PPFDINA": ("PP_F", "Future Prices Paid"),
    "PRFDINA": ("PR_F", "Future Prices Received"),
    "CSFDINA": ("CE_F", "Future Capital Spending"),
    "ASFDINA": ("TS_F", "Future Technology Spending"),
}

# Kansas City: transposed format — row labels map to (std_code, display_name)
# Separate current/future dicts since same row labels appear in both sections.
KC_MFG_CURRENT = {
    "Composite Index": ("GA_C", "Composite Index"),
    "Production": ("PROD_C", "Production"),
    "Volume of shipments": ("SH_C", "Shipments"),
    "Volume of new orders": ("NO_C", "New Orders"),
    "Backlog of orders": ("UO_C", "Unfilled Orders"),
    "Number of employees": ("EMP_C", "Employment"),
    "Average employee workweek": ("AW_C", "Avg Workweek"),
    "Prices received for finished product": ("PR_C", "Prices Received"),
    "Prices paid for raw materials": ("PP_C", "Prices Paid"),
    "Capital expenditures": ("CE_C", "Capital Expenditures"),
    "New orders for exports": ("EX_C", "Export Orders"),
    "Supplier delivery time": ("DT_C", "Supplier Delivery Time"),
    "Inventories: Materials": ("RMI_C", "Inventories: Materials"),
    "Inventories: Finished goods": ("IV_C", "Inventories: Finished Goods"),
}

KC_MFG_FUTURE = {
    "Composite Index": ("GA_F", "Future Composite Index"),
    "Production": ("PROD_F", "Future Production"),
    "Volume of shipments": ("SH_F", "Future Shipments"),
    "Volume of new orders": ("NO_F", "Future New Orders"),
    "Backlog of orders": ("UO_F", "Future Unfilled Orders"),
    "Number of employees": ("EMP_F", "Future Employment"),
    "Average employee workweek": ("AW_F", "Future Avg Workweek"),
    "Prices received for finished product": ("PR_F", "Future Prices Received"),
    "Prices paid for raw materials": ("PP_F", "Future Prices Paid"),
    "Capital expenditures": ("CE_F", "Future Capital Expenditures"),
    "New orders for exports": ("EX_F", "Future Export Orders"),
    "Supplier delivery time": ("DT_F", "Future Supplier Delivery Time"),
    "Inventories: Materials": ("RMI_F", "Future Inventories: Materials"),
    "Inventories: Finished goods": ("IV_F", "Future Inventories: Finished Goods"),
}

KC_SVC_CURRENT = {
    "Composite Index": ("GA_C", "Composite Index"),
    "General Revenue/Sales": ("REV_C", "Revenue/Sales"),
    "Number of Employees": ("EMP_C", "Employment"),
    "Employee Hours Worked": ("AW_C", "Hours Worked"),
    "Part-Time/Temporary Employment": ("PEMP_C", "Part-Time Employment"),
    "Wages and Benefits": ("WB_C", "Wages & Benefits"),
    "Inventory Levels": ("IV_C", "Inventories"),
    "Credit Conditions/Access to Credit": ("CR_C", "Credit Conditions"),
    "Capital Expenditures": ("CE_C", "Capital Expenditures"),
    "Input Prices": ("PP_C", "Prices Paid"),
    "Selling Prices": ("PR_C", "Prices Received"),
}

KC_SVC_FUTURE = {
    "Composite Index": ("GA_F", "Future Composite Index"),
    "General Revenue/Sales": ("REV_F", "Future Revenue/Sales"),
    "Number of Employees": ("EMP_F", "Future Employment"),
    "Employee Hours Worked": ("AW_F", "Future Hours Worked"),
    "Part-Time/Temporary Employment": ("PEMP_F", "Future Part-Time Employment"),
    "Wages and Benefits": ("WB_F", "Future Wages & Benefits"),
    "Inventory Levels": ("IV_F", "Future Inventories"),
    "Credit Conditions/Access to Credit": ("CR_F", "Future Credit Conditions"),
    "Capital Expenditures": ("CE_F", "Future Capital Expenditures"),
    "Input Prices": ("PP_F", "Future Prices Paid"),
    "Selling Prices": ("PR_F", "Future Prices Received"),
}

# Combined for build_series_list name lookups
KC_MFG_COLS = {**{v[0]: v for v in KC_MFG_CURRENT.values()},
               **{v[0]: v for v in KC_MFG_FUTURE.values()}}
KC_SVC_COLS = {**{v[0]: v for v in KC_SVC_CURRENT.values()},
               **{v[0]: v for v in KC_SVC_FUTURE.values()}}

CITY_NAMES = {
    "philly": "Philadelphia",
    "dallas": "Dallas",
    "richmond": "Richmond",
    "ny": "New York",
    "kc": "Kansas City",
}

COL_MAPS = {
    "philly_mfg": PHILLY_MFG_COLS,
    "philly_svc": PHILLY_SVC_COLS,
    "dallas_mfg": DALLAS_MFG_COLS,
    "dallas_svc": DALLAS_SVC_COLS,
    "richmond_mfg": RICHMOND_MFG_COLS,
    "richmond_svc": RICHMOND_SVC_COLS,
    "ny_mfg": NY_MFG_COLS,
    "ny_svc": NY_SVC_COLS,
    "kc_mfg": KC_MFG_COLS,
    "kc_svc": KC_SVC_COLS,
}

# Common components across most surveys (for labeling)
COMMON_MFG = {"GA_C", "NO_C", "SH_C", "EMP_C", "PP_C", "PR_C",
              "GA_F", "NO_F", "SH_F", "EMP_F", "PP_F", "PR_F"}
COMMON_SVC = {"GA_C", "EMP_C", "PP_C", "PR_C", "CE_C",
              "GA_F", "EMP_F", "PP_F", "PR_F", "CE_F"}

# Month abbreviation -> number
MONTH_ABBR = {
    "Jan": "01", "Feb": "02", "Mar": "03", "Apr": "04",
    "May": "05", "Jun": "06", "Jul": "07", "Aug": "08",
    "Sep": "09", "Oct": "10", "Nov": "11", "Dec": "12",
}


def parse_date(raw):
    """Normalize various date formats to YYYY-MM-01."""
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None

    # YYYY-MM-DD (NY Fed, Richmond datetime)
    if len(s) >= 10 and s[4] == '-':
        return s[:7] + "-01"

    # datetime object (from openpyxl)
    if hasattr(raw, 'strftime'):
        return raw.strftime("%Y-%m-01")

    # Mon-YY or Mon-YYYY (Philly, Dallas)
    if '-' in s and len(s) <= 8:
        parts = s.split('-')
        if len(parts) == 2:
            mon, yr = parts
            mm = MONTH_ABBR.get(mon)
            if mm:
                if len(yr) == 2:
                    yr = "20" + yr if int(yr) < 50 else "19" + yr
                return f"{yr}-{mm}-01"

    return None


def parse_value(raw):
    """Parse a cell value to float or None."""
    if raw is None:
        return None
    s = str(raw).strip()
    if not s or s in ("ND", "#N/A", "NA", "N/A", "."):
        return None
    try:
        return round(float(s), 2)
    except (ValueError, TypeError):
        return None


def download(key):
    """Download a source file and return raw bytes."""
    src = SOURCES[key]
    print(f"  Downloading {key}...")
    resp = retry_request(src["url"])
    print(f"    {len(resp.content):,} bytes")
    return resp.content


def parse_csv_source(content, col_map, date_col):
    """Parse a CSV file into series dict. Returns {std_code: [{"date","value"}]}."""
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    headers = reader.fieldnames

    # Build col index: source_col -> (std_code, display_name)
    active_cols = {}
    for h in headers:
        if h in col_map:
            active_cols[h] = col_map[h]

    series = {code: [] for code, _ in active_cols.values()}

    for row in reader:
        date = parse_date(row.get(date_col))
        if not date:
            continue
        for src_col, (std_code, _) in active_cols.items():
            val = parse_value(row.get(src_col))
            series[std_code].append({"date": date, "value": val})

    return series


def parse_xlsx_source(content, col_map, sheet_name=None):
    """Parse an Excel file into series dict."""
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return {}

    headers = [str(h).strip() if h else "" for h in rows[0]]

    # Build col index
    active_cols = {}
    for i, h in enumerate(headers):
        if h in col_map:
            active_cols[i] = col_map[h]

    series = {code: [] for code, _ in active_cols.values()}

    for row in rows[1:]:
        date = parse_date(row[0])
        if not date:
            continue
        for col_idx, (std_code, _) in active_cols.items():
            if col_idx < len(row):
                val = parse_value(row[col_idx])
                series[std_code].append({"date": date, "value": val})

    return series


def parse_kc_xlsx(content, current_map, future_map, sheet_name=None):
    """Parse a Kansas City Fed Excel file (transposed: dates in columns, series in rows).

    The file has multiple sections. We want the seasonally adjusted sections:
      - 'Versus a Month Ago' + '(seasonally adjusted)' -> current
      - 'Expected in Six Months' + '(seasonally adjusted)' -> future
    """
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    # Row 3 contains dates across columns (col 2 onwards)
    dates = []
    for c in range(2, ws.max_column + 1):
        raw = ws.cell(3, c).value
        d = parse_date(raw)
        if d:
            dates.append((c, d))

    if not dates:
        return {}

    # Walk rows to find SA sections
    series = {}
    section = None  # "current_sa" or "future_sa" or None

    for r in range(4, ws.max_row + 1):
        label = ws.cell(r, 1).value
        if label is None:
            continue
        label = str(label).strip()

        # Detect section headers
        if label == "Versus a Month Ago":
            # Check next row for (seasonally adjusted)
            next_label = ws.cell(r + 1, 1).value
            if next_label and "seasonally adjusted" in str(next_label).lower():
                section = "current_sa"
            else:
                section = None
            continue
        elif label == "Expected in Six Months":
            next_label = ws.cell(r + 1, 1).value
            if next_label and "seasonally adjusted" in str(next_label).lower():
                section = "future_sa"
            else:
                section = None
            continue
        elif label == "Versus a Year Ago":
            section = None
            continue
        elif "seasonally adjusted" in label.lower():
            # This is the "(seasonally adjusted)" or "(not seasonally adjusted)" tag
            if "not" in label.lower():
                section = None
            continue

        if section is None:
            continue

        # Map row label to standardized code
        col_map = current_map if section == "current_sa" else future_map
        if label not in col_map:
            continue

        std_code, _ = col_map[label]

        data_points = []
        for col_idx, date_str in dates:
            val = parse_value(ws.cell(r, col_idx).value)
            data_points.append({"date": date_str, "value": val})

        series[std_code] = data_points

    return series


def build_series_list(parsed, city_key, col_map, survey_type):
    """Convert parsed series dict to the site's standard series list format."""
    city_name = CITY_NAMES[city_key]
    common = COMMON_MFG if survey_type == "mfg" else COMMON_SVC

    # Get display name lookup from col_map values
    name_lookup = {code: name for code, name in col_map.values()}

    series_list = []
    order = 0
    for std_code, data_points in parsed.items():
        # Filter out entries with no valid data
        valid = [d for d in data_points if d["value"] is not None]
        if len(valid) < 2:
            continue

        display_name = name_lookup.get(std_code, std_code)
        series_id = f"{city_key.upper()}_{std_code}"

        series_list.append({
            "id": series_id,
            "name": f"{city_name} - {display_name}",
            "display_order": order,
            "data": sorted(data_points, key=lambda d: d["date"]),
        })
        order += 1

    return series_list


def fetch_and_parse(key):
    """Download and parse a single source into a series list."""
    src = SOURCES[key]
    content = download(key)
    col_map = COL_MAPS[key]

    city = key.split("_")[0]
    survey_type = key.split("_")[1]

    if src["format"] == "csv":
        # Determine date column
        if key.startswith("philly"):
            date_col = "DATE"
        else:
            date_col = "surveyDate"
        parsed = parse_csv_source(content, col_map, date_col)
    elif src["format"] == "kc_xlsx":
        # Kansas City transposed format
        if key == "kc_mfg":
            parsed = parse_kc_xlsx(content, KC_MFG_CURRENT, KC_MFG_FUTURE)
        else:
            parsed = parse_kc_xlsx(content, KC_SVC_CURRENT, KC_SVC_FUTURE)
    else:
        # Excel
        sheet = None
        if key == "philly_svc":
            sheet = "Diffusion"
        parsed = parse_xlsx_source(content, col_map, sheet_name=sheet)

    series = build_series_list(parsed, city, col_map, survey_type)
    return series


def run():
    print("Fetching Federal Reserve regional survey data...")

    mfg_series = []
    svc_series = []

    for key in SOURCES:
        try:
            series = fetch_and_parse(key)
            if "_mfg" in key:
                mfg_series.extend(series)
            else:
                svc_series.extend(series)
            print(f"    {len(series)} series parsed")
        except Exception as e:
            print(f"    ERROR: {e}")

    # Reassign display_order across all cities
    for i, s in enumerate(mfg_series):
        s["display_order"] = i
    for i, s in enumerate(svc_series):
        s["display_order"] = i

    now = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

    if mfg_series:
        write_json({
            "metadata": {
                "title": "Fed Regional Surveys - Manufacturing",
                "source": "Federal Reserve Banks of Philadelphia, Dallas, Richmond, New York, Kansas City",
                "unit": "Diffusion Index",
                "frequency": "monthly",
                "last_updated": now,
            },
            "series": mfg_series,
        }, "fed_surveys/fed_mfg.json")
        print(f"  Manufacturing: {len(mfg_series)} total series")

    if svc_series:
        write_json({
            "metadata": {
                "title": "Fed Regional Surveys - Services",
                "source": "Federal Reserve Banks of Philadelphia, Dallas, Richmond, New York, Kansas City",
                "unit": "Diffusion Index",
                "frequency": "monthly",
                "last_updated": now,
            },
            "series": svc_series,
        }, "fed_surveys/fed_svc.json")
        print(f"  Services: {len(svc_series)} total series")


if __name__ == "__main__":
    run()
